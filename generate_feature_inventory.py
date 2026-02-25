import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BookEx Feature Inventory"

# ── Column definitions ──────────────────────────────────────────────────────
headers = [
    "Module Name",
    "Feature Name",
    "Short Description",
    "User Roles Involved",
    "Backend Required",
    "Real-Time Involved",
    "AI Involved",
    "Priority",
]

# ── Feature data ────────────────────────────────────────────────────────────
# Only CONFIRMED IMPLEMENTED features (audited against actual codebase)
# (Module, Feature, Description, Roles, Backend, RealTime, AI, Priority)
features = [

    # ===========================================================
    # AUTHENTICATION MODULE
    # Confirmed: signUpUser, requestPasswordReset, resetPassword,
    # CredentialsProvider (bcrypt), NextAuth JWT, profile-completion-guard,
    # auth-rate-limiting.ts, middleware.ts, RBAC via role field
    # NOT implemented: Google OAuth (only CredentialsProvider in auth-config.ts)
    # ===========================================================
    ("Authentication", "User Registration (Email/Password)",
     "Registers a new user with email, hashed password, and default role assignment",
     "Guest", "Yes", "No", "No", "High"),

    ("Authentication", "Email Uniqueness Validation",
     "Checks MongoDB for duplicate email before completing registration",
     "Guest", "Yes", "No", "No", "High"),

    ("Authentication", "Password Hashing (bcrypt)",
     "Hashes plain-text passwords using bcryptjs before persisting to database",
     "System", "Yes", "No", "No", "High"),

    ("Authentication", "User Login (Credentials)",
     "Authenticates user via email and password, creates NextAuth JWT session",
     "Guest", "Yes", "No", "No", "High"),

    ("Authentication", "Session Token Management",
     "Issues and rotates JWT tokens via NextAuth; role, status, profileCompleted in token",
     "All Authenticated Users", "Yes", "No", "No", "High"),

    ("Authentication", "Logout / Session Destruction",
     "Invalidates current session token and clears client-side auth state",
     "All Authenticated Users", "Yes", "No", "No", "High"),

    ("Authentication", "CSRF Protection",
     "Next.js server actions enforce CSRF via built-in origin checking",
     "System", "Yes", "No", "No", "High"),

    ("Authentication", "Role-Based Access Control (RBAC)",
     "Enforces platform-level roles (user / admin) on routes and server actions",
     "System", "Yes", "No", "No", "High"),

    ("Authentication", "Profile Completion Enforcement",
     "ProfileCompletionGuard component redirects users to complete mandatory profile fields",
     "User", "Yes", "No", "No", "High"),

    ("Authentication", "Account Status Gate",
     "Blocks login for suspended accounts with descriptive error in NextAuth authorize",
     "System", "Yes", "No", "No", "High"),

    ("Authentication", "Auth Rate Limiting",
     "checkAuthRateLimit / recordAuthResult cap login attempts per IP using Redis",
     "System", "Yes", "No", "No", "High"),

    ("Authentication", "Password Reset Request",
     "requestPasswordReset generates a time-limited token and sends reset email",
     "Guest", "Yes", "No", "No", "High"),

    ("Authentication", "Password Reset Confirmation",
     "resetPassword validates token, updates bcrypt hash, and invalidates the token",
     "Guest", "Yes", "No", "No", "High"),

    ("Authentication", "Middleware Route Protection",
     "Next.js middleware intercepts unauthenticated requests and redirects to login",
     "System", "Yes", "No", "No", "High"),

    # ===========================================================
    # USER PROFILE MODULE
    # Confirmed: updateUserProfile, getUserForUpdate, file-storage upload,
    # getUserReviews, getUserExchanges, completeUserProfile, checkProfileCompletion,
    # getUserEmailPreferences, updateEmailPreferences, cities-database
    # NOT implemented: Update Password while authenticated (no server action found),
    #                  Account Soft Delete by self (admin-only deleteUser)
    # ===========================================================
    ("User Profile", "View Own Profile",
     "Displays user's own profile including bio, ratings, city, and exchange history",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "View Other User's Profile",
     "Displays a public-facing profile page for any non-deleted, active user",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "Edit Profile Details",
     "updateUserProfile allows user to update display name, username, city",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "Upload Profile Image",
     "Handles profile photo upload via file-storage.ts, updates avatarUrl on profile",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "Profile Image Validation",
     "Enforces file type and size limits on profile photo uploads server-side",
     "System", "Yes", "No", "No", "Medium"),

    ("User Profile", "View Received Ratings",
     "getUserReviews returns aggregated rating score and individual reviews",
     "User", "Yes", "No", "No", "Medium"),

    ("User Profile", "View Exchange History",
     "getUserExchanges lists all past and active exchanges with status and partner",
     "User", "Yes", "No", "No", "Medium"),

    ("User Profile", "Account Reactivation (Admin)",
     "activateUser admin action reactivates a suspended account",
     "Admin", "Yes", "No", "No", "Medium"),

    ("User Profile", "Profile Completion Modal",
     "Profile completion modal and progress guard enforced on first login",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "City / Location Setting",
     "User sets city from validated cities-database; used for geo-relevance in book search",
     "User", "Yes", "No", "No", "High"),

    ("User Profile", "Email Notification Preferences",
     "getUserEmailPreferences / updateEmailPreferences toggle per-category email opt-in",
     "User", "Yes", "No", "No", "Medium"),

    ("User Profile", "Account Status Display",
     "User profile shows current account status (active / suspended)",
     "User", "Yes", "No", "No", "Medium"),

    # ===========================================================
    # BOOK MODULE
    # Confirmed: listBook, updateBookListing, deleteBookListing, updateBookStatus,
    # renewBookListing, getBookForEdit, rate-limiting RATE_LIMITS,
    # file-storage image upload, content-moderation-middleware,
    # content-moderation-dashboard, AI flows (summary, search, recommendations),
    # database-indexes text index on title/author/genre
    # NOT implemented: Admin book approval queue (no approveBook action found),
    #                  "Mark Book as Unavailable" handled by updateBookStatus
    # ===========================================================
    ("Book", "Create Book Listing",
     "listBook submits new book with title, author, genre, condition, description, image",
     "User", "Yes", "No", "No", "High"),

    ("Book", "Edit Book Listing",
     "updateBookListing allows owner to update any field of an existing listing",
     "User", "Yes", "No", "No", "High"),

    ("Book", "Delete Book Listing (Soft)",
     "deleteBookListing sets deletedAt timestamp; listing hidden from all queries",
     "User", "Yes", "No", "No", "High"),

    ("Book", "Book Listing Rate Limiting",
     "checkUserRateLimit caps new listings at 10 per day per user via Redis",
     "System", "Yes", "No", "No", "High"),

    ("Book", "Book Image Upload",
     "Validates image type/size then stores via file-storage.ts; URL saved on listing",
     "User", "Yes", "No", "No", "High"),

    ("Book", "Book Status Transitions",
     "updateBookStatus manages state machine: available/in-exchange/exchanged/unavailable",
     "System", "Yes", "No", "No", "High"),

    ("Book", "Book Listing Renewal",
     "renewBookListing refreshes listing timestamp to keep it active",
     "User", "Yes", "No", "No", "Medium"),

    ("Book", "Search Books (Text)",
     "Full-text search on title, author, genre fields using MongoDB text index",
     "User / Guest", "Yes", "No", "No", "High"),

    ("Book", "Filter Books (Genre/Condition/City)",
     "Applies compound filters: genre, condition, city, and availability status",
     "User / Guest", "Yes", "No", "No", "High"),

    ("Book", "Book Listing Pagination",
     "Returns paginated results with cursor/offset to handle large datasets efficiently",
     "User / Guest", "Yes", "No", "No", "High"),

    ("Book", "AI Book Summary Generation",
     "generateBookSummary Genkit flow uses Gemini to create a short book summary",
     "User", "Yes", "No", "Yes", "Medium"),

    ("Book", "AI Book Condition Analysis",
     "analyzeBookCondition flow uses Gemini vision to suggest condition and price from photo",
     "User", "Yes", "No", "Yes", "Medium"),

    ("Book", "AI Book Recommendations",
     "getBookRecommendations Genkit flow suggests books based on exchange history",
     "User", "Yes", "No", "Yes", "Medium"),

    ("Book", "AI-Assisted Book Search",
     "aiAssistantFlow parses natural language queries and maps them to structured filters",
     "User", "Yes", "No", "Yes", "Medium"),

    ("Book", "Book Content Moderation (AI)",
     "content-moderation-middleware scans new listings for policy violations before publish",
     "System", "Yes", "No", "Yes", "High"),

    ("Book", "Admin Content Removal",
     "removeContentAndResolveReport allows admin to remove flagged book listings",
     "Admin", "Yes", "No", "No", "High"),

    ("Book", "View Book Detail",
     "Full book info page including owner profile, condition, and exchange/donate CTA",
     "User / Guest", "Yes", "No", "No", "High"),

    ("Book", "My Listings View",
     "Displays all books uploaded by authenticated user with status indicators",
     "User", "Yes", "No", "No", "High"),

    ("Book", "Book Condition Validation",
     "Validates condition field against enum: new / like-new / used / worn",
     "System", "Yes", "No", "No", "Medium"),

    ("Book", "Book Ownership Verification",
     "validateResourceAccess verifies only book owner or admin can mutate listing",
     "System", "Yes", "No", "No", "High"),

    ("Book", "MongoDB Text Index on Books",
     "Compound text index on title + author + genre created via database-setup.ts",
     "System", "Yes", "No", "No", "High"),

    # ===========================================================
    # EXCHANGE MODULE
    # Confirmed: proposeExchange, acceptExchange, confirmExchangeCompletion,
    # cancelExchange, getUserExchanges, startExchangeChat, submitReview,
    # notification-utils exchange status notifications, email notifications
    # NOT implemented: rejectExchange as a distinct action
    #   (only cancelExchange + acceptExchange found; 'rejected' in enum but no handler)
    # ===========================================================
    ("Exchange", "Propose Exchange",
     "proposeExchange initiates proposal with offered book and requested book",
     "User", "Yes", "No", "No", "High"),

    ("Exchange", "Accept Exchange Proposal",
     "acceptExchange transitions both books to in-exchange status",
     "User", "Yes", "No", "No", "High"),

    ("Exchange", "Cancel Exchange",
     "cancelExchange allows both requester and owner to cancel; rolls back book statuses",
     "User", "Yes", "No", "No", "High"),

    ("Exchange", "Complete Exchange",
     "confirmExchangeCompletion marks both books as exchanged and triggers rating prompt",
     "User", "Yes", "No", "No", "High"),

    ("Exchange", "Exchange State Machine",
     "Enforces valid status transitions (proposed → accepted → completed / cancelled)",
     "System", "Yes", "No", "No", "High"),

    ("Exchange", "Exchange Chat Integration",
     "startExchangeChat auto-creates dedicated chat room when exchange is accepted",
     "System", "Yes", "Yes", "No", "High"),

    ("Exchange", "Exchange Real-Time Status Updates",
     "Socket.IO pushes live exchange status events to both participants",
     "User", "Yes", "Yes", "No", "High"),

    ("Exchange", "Exchange History View",
     "getUserExchanges lists all past and active exchanges with status and dates",
     "User", "Yes", "No", "No", "Medium"),

    ("Exchange", "Duplicate Proposal Prevention",
     "Checks for existing pending proposal between same users and books before creating",
     "System", "Yes", "No", "No", "High"),

    ("Exchange", "Exchange Eligibility Check",
     "Verifies both books are available and both users are active before proposal",
     "System", "Yes", "No", "No", "High"),

    ("Exchange", "Post-Exchange Rating & Review",
     "submitReview lets each party rate the other (1-5 stars) with optional comment",
     "User", "Yes", "No", "No", "High"),

    ("Exchange", "Exchange Transaction Handling",
     "MongoDB multi-document transaction wraps book status + exchange record creation",
     "System", "Yes", "No", "No", "High"),

    ("Exchange", "Exchange Email Notifications",
     "sendExchangeProposalEmail and sendExchangeStatusUpdateEmail notify participants",
     "System", "Yes", "No", "No", "High"),

    # ===========================================================
    # COMMUNITY MODULE
    # Confirmed via actions.ts + community-admin-actions.ts:
    # createCommunity, updateCommunitySettings (includes privacy toggle), toggleCommunityMembership,
    # requestToJoinCommunity, approveJoinRequest, rejectJoinRequest, removeMemberFromCommunity,
    # banMemberFromCommunity, unbanMemberFromCommunity, promoteMember, demoteMember,
    # transferCommunityOwnership, createPost, editPost, deletePost, addComment, adminDeleteComment,
    # togglePostLike, setPinPost, setLockPost, searchCommunities, reportCommunityContent,
    # chat-channel.tsx (real-time channel chat), community-page community list
    # ===========================================================
    ("Community", "Create Community",
     "createCommunity creates community with name, description, image, and visibility",
     "User", "Yes", "No", "No", "High"),

    ("Community", "Edit Community Details",
     "updateCommunitySettings allows owner/admin to update name, description, rules",
     "Community Owner / Admin", "Yes", "No", "No", "High"),

    ("Community", "Delete Community (Admin)",
     "Admin can delete community via /api/admin/communities/[communityId] route",
     "Admin", "Yes", "No", "No", "High"),

    ("Community", "Community Privacy Toggle (Public/Private)",
     "admin-settings-panel toggles visibility between public and private via updateCommunitySettings",
     "Community Owner / Admin", "Yes", "No", "No", "High"),

    ("Community", "Join Community (Public)",
     "toggleCommunityMembership immediately adds member to public community",
     "User", "Yes", "No", "No", "High"),

    ("Community", "Request to Join Community (Private)",
     "requestToJoinCommunity adds user to pendingRequests of a private community",
     "User", "Yes", "No", "No", "High"),

    ("Community", "Approve Join Request",
     "approveJoinRequest grants Member role and notifies requesting user",
     "Community Owner / Admin / Moderator", "Yes", "Yes", "No", "High"),

    ("Community", "Reject Join Request",
     "rejectJoinRequest removes pending request and notifies user",
     "Community Owner / Admin / Moderator", "Yes", "No", "No", "High"),

    ("Community", "Leave Community",
     "toggleCommunityMembership(isMember=true) removes user from members list",
     "User", "Yes", "No", "No", "High"),

    ("Community", "Promote Member to Moderator",
     "promoteMember / promoteMember-in-actions elevates member to moderator role",
     "Community Owner / Admin", "Yes", "No", "No", "High"),

    ("Community", "Demote Moderator to Member",
     "demoteMember reduces a moderator back to regular member role",
     "Community Owner / Admin", "Yes", "No", "No", "Medium"),

    ("Community", "Ownership Transfer",
     "transferCommunityOwnership reassigns ownership to another active member",
     "Community Owner", "Yes", "No", "No", "Medium"),

    ("Community", "Ban Member",
     "banMemberFromCommunity removes member and records ban; blocks future joins",
     "Community Owner / Admin", "Yes", "No", "No", "High"),

    ("Community", "Unban Member",
     "unbanMemberFromCommunity lifts ban to allow re-joining",
     "Community Owner / Admin", "Yes", "No", "No", "Medium"),

    ("Community", "Remove Member",
     "removeMemberFromCommunity kicks a member without a ban record",
     "Community Owner / Admin / Moderator", "Yes", "No", "No", "Medium"),

    ("Community", "Pending Requests Panel",
     "getPendingJoinRequests returns all pending join requests for admin review",
     "Community Admin / Owner", "Yes", "No", "No", "High"),

    ("Community", "Create Community Post",
     "createPost publishes a text post to the community feed",
     "Community Member", "Yes", "No", "No", "High"),

    ("Community", "Edit Community Post",
     "editPost updates post content; restricted to post author or moderator",
     "Post Author / Moderator", "Yes", "No", "No", "High"),

    ("Community", "Delete Community Post (Soft)",
     "deletePost / adminDeletePost soft-deletes post by setting deletedAt",
     "Post Author / Moderator / Admin", "Yes", "No", "No", "High"),

    ("Community", "Comment on Post",
     "addComment adds a comment (with optional parentId for threading) to a post",
     "Community Member", "Yes", "No", "No", "High"),

    ("Community", "Delete Comment (Soft)",
     "adminDeleteComment soft-deletes a comment from a post",
     "Comment Author / Moderator", "Yes", "No", "No", "Medium"),

    ("Community", "Like / Unlike Post",
     "togglePostLike toggles a like reaction and updates like count",
     "Community Member", "Yes", "No", "No", "Medium"),

    ("Community", "Pin Post",
     "setPinPost pins/unpins a post to the top of the community feed",
     "Community Admin / Moderator", "Yes", "No", "No", "Medium"),

    ("Community", "Lock Post",
     "setLockPost prevents further comments on a post",
     "Community Admin / Moderator", "Yes", "No", "No", "Medium"),

    ("Community", "Community Chat Channel",
     "chat-channel.tsx provides real-time Socket.IO channel chat per community",
     "Community Member", "Yes", "Yes", "No", "High"),

    ("Community", "Create Community Channel",
     "createChannel server action adds a new forum or chat channel to a community",
     "Community Owner / Admin", "Yes", "No", "No", "Medium"),

    ("Community", "Community Search",
     "searchCommunities full-text search communities by name",
     "User / Guest", "Yes", "No", "No", "Medium"),

    ("Community", "Community Discovery / Browse",
     "Community list page shows all public communities with member count",
     "User / Guest", "Yes", "No", "No", "Medium"),

    ("Community", "Report Community Content",
     "reportCommunityContent lets members report posts/comments to admin",
     "Community Member", "Yes", "No", "No", "High"),

    ("Community", "Community Moderation Log",
     "getCommunityModerationLog returns paginated moderation actions for a community",
     "Community Owner / Admin", "Yes", "No", "No", "Medium"),

    # ===========================================================
    # MESSAGING MODULE
    # Confirmed: startChat, startExchangeChat, getUserChats, getChatDetails,
    # Socket.IO message events, typing events, presenceManager, readAt marking,
    # file-attachment-upload.tsx, message-attachment-display.tsx,
    # /api/messages/chats/[chatId]/read route, unreadCount in messages-page.tsx,
    # /api/messages/chats/[chatId]/delete (soft-deletes conversation for user)
    # NOT implemented: individual message deletion (no per-message delete endpoint found)
    # ===========================================================
    ("Messaging", "Direct Message (DM) Initiation",
     "startChat creates or retrieves a DM conversation between two users",
     "User", "Yes", "Yes", "No", "High"),

    ("Messaging", "Send Text Message",
     "Socket.IO sendMessage event persists message to MongoDB and broadcasts to room",
     "User", "Yes", "Yes", "No", "High"),

    ("Messaging", "Message Persistence",
     "All chat messages stored in MongoDB with chatId, senderId, content, timestamp",
     "System", "Yes", "No", "No", "High"),

    ("Messaging", "Chat Room Join via Socket.IO",
     "Client emits joinRoom on page load; server validates membership before room entry",
     "System", "No", "Yes", "No", "High"),

    ("Messaging", "Typing Indicator",
     "Socket.IO typing event emits isTyping to room; receiver shows indicator",
     "User", "No", "Yes", "No", "Medium"),

    ("Messaging", "Read Receipts",
     "/api/messages/chats/[chatId]/read updates messages.$[elem].readAt in MongoDB",
     "User", "Yes", "Yes", "No", "Medium"),

    ("Messaging", "File / Image Attachment",
     "file-attachment-upload.tsx uploads file; message stores attachment URL reference",
     "User", "Yes", "Yes", "No", "Medium"),

    ("Messaging", "Attachment Storage",
     "Files uploaded via /api/files endpoint; stored and referenced by URL in message",
     "System", "Yes", "No", "No", "Medium"),

    ("Messaging", "Chat Message History Load",
     "getChatDetails retrieves message history with pagination on chat open",
     "User", "Yes", "No", "No", "High"),

    ("Messaging", "Chat Room Access Control",
     "Server validates participant membership before allowing message send or read",
     "System", "Yes", "No", "No", "High"),

    ("Messaging", "Exchange Chat Room Auto-Creation",
     "startExchangeChat auto-creates dedicated chat room upon exchange acceptance",
     "System", "Yes", "Yes", "No", "High"),

    ("Messaging", "Community Chat Room Scoping",
     "Community channels scoped to members via /api/communities/[communityId]/channels",
     "Community Member", "Yes", "Yes", "No", "High"),

    ("Messaging", "Unread Message Count",
     "unreadCount tracked per chat in messages-page.tsx; badge shown in conversation list",
     "User", "Yes", "No", "No", "Medium"),

    ("Messaging", "Conversation List",
     "getUserChats retrieves all DM and exchange chats sorted by last message",
     "User", "Yes", "No", "No", "High"),

    ("Messaging", "Delete Conversation (Soft)",
     "/api/messages/chats/[chatId]/delete soft-deletes conversation for current user",
     "User", "Yes", "No", "No", "Medium"),

    ("Messaging", "Online Presence Indicator",
     "presenceManager.setUserOnline + Socket.IO presenceUpdate broadcast online status",
     "User", "No", "Yes", "No", "Low"),

    # ===========================================================
    # NOTIFICATION MODULE
    # Confirmed: getUserNotifications, getNotifications, markNotificationsAsRead,
    # markNotificationAsRead, deleteNotification, getUnreadNotificationCount,
    # notification-utils (createExchangeStatusNotification etc.),
    # email.ts sendExchangeProposalEmail, sendExchangeStatusUpdateEmail,
    # admin-notifications.ts, notification-provider.tsx, use-admin-notifications.ts
    # NOT implemented: Mention Notification (no @mention system found anywhere)
    #                  Post Review Outcome Notification (no book approval queue)
    # ===========================================================
    ("Notifications", "New Message Notification",
     "Socket.IO newMessage event triggers in-app notification for offline recipient",
     "User", "Yes", "Yes", "No", "High"),

    ("Notifications", "Exchange State Change Notification",
     "createExchangeStatusNotification persists and delivers in-app exchange updates",
     "User", "Yes", "Yes", "No", "High"),

    ("Notifications", "New Exchange Proposal Notification",
     "proposeExchange creates notification for book owner on new incoming proposal",
     "User", "Yes", "Yes", "No", "High"),

    ("Notifications", "Community Join Request Notification",
     "requestToJoinCommunity notifies community admin of new pending join request",
     "Community Admin / Owner", "Yes", "Yes", "No", "High"),

    ("Notifications", "Join Request Approved/Rejected Notification",
     "approveJoinRequest / rejectJoinRequest sends outcome notification to requester",
     "User", "Yes", "Yes", "No", "High"),

    ("Notifications", "Community Post Notification",
     "New post creation triggers notification to community members via notification-utils",
     "Community Member", "Yes", "Yes", "No", "Medium"),

    ("Notifications", "Mark Notification as Read",
     "markNotificationAsRead updates single notification read state in DB",
     "User", "Yes", "No", "No", "High"),

    ("Notifications", "Mark All Notifications as Read",
     "markNotificationsAsRead bulk-marks all user notifications as read",
     "User", "Yes", "No", "No", "Medium"),

    ("Notifications", "Notification Count Badge",
     "getUnreadNotificationCount / notification-provider updates header badge in real-time",
     "User", "Yes", "Yes", "No", "High"),

    ("Notifications", "Notification List Pagination",
     "getUserNotifications returns paginated notification list sorted by recency",
     "User", "Yes", "No", "No", "Medium"),

    ("Notifications", "Delete Notification",
     "deleteNotification removes a specific notification from user's list",
     "User", "Yes", "No", "No", "Low"),

    ("Notifications", "Email Notification Dispatch",
     "email.ts sends transactional emails for exchanges, proposals, and account events",
     "System", "Yes", "No", "No", "Medium"),

    ("Notifications", "Account Suspension Notification",
     "suspendUser triggers notification and email to affected user on suspension",
     "User", "Yes", "No", "No", "High"),

    ("Notifications", "Admin Alert Notification",
     "admin-notifications.ts pushes alerts to admin on new reports and moderation flags",
     "Admin", "Yes", "Yes", "No", "High"),

    # ===========================================================
    # ADMIN DASHBOARD MODULE
    # Confirmed tabs: overview, users, reports, organizations, security, moderation, database
    # user actions: suspendUser, activateUser, deleteUser
    # report actions: resolveReport, resolveReportWithNotes, bulkResolveReports,
    #                 removeContentAndResolveReport
    # community admin: /api/admin/communities/[communityId]
    # security: security-admin-dashboard.tsx (threats, blocked, contentModeration)
    # content-moderation-dashboard.tsx shows flaggedContent, autoRejected, pendingReview
    # NOT implemented: Dedicated Audit Log Viewer UI, Filter Audit Logs UI,
    #                  View All Book Listings (admin), View Exchange Overview (admin),
    #                  Approve Book Listing (no approval queue)
    # ===========================================================
    ("Admin Dashboard", "View All Users",
     "Admin dashboard users tab shows paginated list with status, role, join date",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Search Users",
     "Admin searches users by name or email from the users tab",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Suspend User Account",
     "suspendUser sets user status to suspended; account immediately blocked",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Activate / Reinstate User Account",
     "activateUser sets user status back to active, restoring platform access",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Hard Delete User Account",
     "deleteUser permanently removes user, books, reviews, and messages from DB",
     "Admin / SuperAdmin", "Yes", "No", "No", "Medium"),

    ("Admin Dashboard", "Remove Book / Content via Report",
     "removeContentAndResolveReport removes flagged book, post, or comment and closes report",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "View All Reports",
     "getAdminReports returns paginated reports filterable by status and content type",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Resolve Report",
     "resolveReport / resolveReportWithNotes marks report resolved with optional notes",
     "Admin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Dismiss Report",
     "admin-report-actions.tsx dismisses report with no action via resolveReport",
     "Admin", "Yes", "No", "No", "Medium"),

    ("Admin Dashboard", "Bulk Resolve Reports",
     "bulkResolveReports resolves multiple reports in one action",
     "Admin", "Yes", "No", "No", "Medium"),

    ("Admin Dashboard", "System Metrics Dashboard",
     "Comprehensive dashboard shows total users, books, exchanges, and security KPIs",
     "Admin / SuperAdmin", "Yes", "No", "No", "Medium"),

    ("Admin Dashboard", "Manage Communities (Admin)",
     "community-admin-dashboard.tsx lets admin view and delete communities",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Admin Role Assignment",
     "First-user becomes admin; admin can suspend/delete other users",
     "SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "View AI / Security Moderation Flags",
     "security-admin-dashboard.tsx displays flaggedContent, autoRejected, pendingReview stats",
     "Admin", "Yes", "No", "Yes", "High"),

    ("Admin Dashboard", "Security Monitoring Dashboard",
     "SecurityAdminDashboard shows threats, blocked IPs, health score, and alert resolution",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Manage Organizations",
     "Organizations tab supports approve, reject, edit, and delete organization records",
     "Admin / SuperAdmin", "Yes", "No", "No", "High"),

    ("Admin Dashboard", "Database Management Panel",
     "DatabaseManagement tab allows index creation, health check, and maintenance runs",
     "Admin / SuperAdmin", "Yes", "No", "No", "Medium"),

    # ===========================================================
    # AI MODULE
    # Confirmed AI flows in src/ai/flows/:
    # - generate-book-summary.ts (generateBookSummary)
    # - get-book-recommendations.ts (getBookRecommendations)
    # - intelligent-book-search.ts (aiAssistantFlow)
    # - analyze-book-condition.ts (analyzeBookCondition)
    # Supporting: ai-validation.ts, rate-limiter.ts, content-moderation.ts
    # NOT implemented: Separate "AI Description Generation" flow (summary covers this)
    #                  "AI Audit Logging" (logs via activity-logging, not AI-specific)
    # ===========================================================
    ("AI", "AI Book Summary Generation (Genkit Flow)",
     "generateBookSummary Genkit flow calls Gemini to generate a short book summary",
     "User", "Yes", "No", "Yes", "Medium"),

    ("AI", "AI Book Condition Analysis (Genkit Flow)",
     "analyzeBookCondition uses Gemini vision to suggest condition enum and price from photo",
     "User", "Yes", "No", "Yes", "Medium"),

    ("AI", "AI Book Recommendation Flow (Genkit)",
     "getBookRecommendations analyses exchange history to suggest relevant books via Gemini",
     "User", "Yes", "No", "Yes", "Medium"),

    ("AI", "AI Natural Language Search Flow (Genkit)",
     "aiAssistantFlow maps free-text queries to structured MongoDB filters via Gemini",
     "User", "Yes", "No", "Yes", "Medium"),

    ("AI", "AI Content Moderation (Rule + AI Hybrid)",
     "ContentModerationSystem in content-moderation.ts validates book content before publish",
     "System", "Yes", "No", "Yes", "High"),

    ("AI", "AI Rate Limiting (per user/hour)",
     "withRateLimit wrapper in rate-limiter.ts caps AI API calls per user with Redis TTL",
     "System", "Yes", "No", "Yes", "High"),

    ("AI", "AI Input Sanitization",
     "ai-validation.ts / checkContentPolicy strips and validates input before Gemini prompt",
     "System", "Yes", "No", "Yes", "High"),

    ("AI", "AI Response Validation",
     "Zod schemas validate Gemini API response shape before returning to client",
     "System", "Yes", "No", "Yes", "High"),

    ("AI", "AI Error Handling & Fallback",
     "Genkit flows catch Gemini errors and return safe fallback responses",
     "System", "Yes", "No", "Yes", "High"),

    # ===========================================================
    # SECURITY & INFRASTRUCTURE MODULE
    # Confirmed:
    # rate-limiting.ts + rate-limiter.ts (Redis), redis-cache.ts,
    # activity-logging.ts (logActivity, detectSuspiciousActivity),
    # soft-delete pattern across all entities (deletedAt),
    # database-setup.ts + database-indexes.ts (compound + text indexes),
    # socket-server.ts (room management + auth middleware),
    # file-storage.ts (S3-compatible), schemas.ts (Zod validation),
    # env-validation.ts, error-handling.ts, logger.ts, security.ts,
    # business-logic-middleware.ts, resource-authorization.ts, CORS in API routes
    # ===========================================================
    ("Security & Infrastructure", "Global API Rate Limiting",
     "rate-limiting.ts + Redis caps requests per IP across all API routes",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Per-User Action Rate Limiting",
     "checkUserRateLimit in rate-limiting.ts enforces per-user limits on high-risk actions",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Redis Cache Layer",
     "redis-cache.ts caches organization lists and other hot data with TTL",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Activity / Audit Logging",
     "logActivity in activity-logging.ts appends structured entries to activity_logs collection",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Suspicious Activity Detection",
     "detectSuspiciousActivity in activity-logging.ts flags anomalous patterns",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Soft Delete Pattern Enforcement",
     "All entities use deletedAt timestamp; queries consistently filter soft-deleted records",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "MongoDB Compound Indexing",
     "database-setup.ts creates compound indexes on userId+status+createdAt for performance",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "MongoDB Text Search Indexing",
     "Text index on book title + author + genre created via database-setup.ts",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Socket.IO Room Management",
     "socket-server.ts maintains isolated rooms per exchange/community with access control",
     "System", "No", "Yes", "No", "High"),

    ("Security & Infrastructure", "Socket.IO Authentication Middleware",
     "Socket.IO handshake middleware verifies NextAuth session before room entry",
     "System", "No", "Yes", "No", "High"),

    ("Security & Infrastructure", "File Upload Validation",
     "validateFileFromFormData checks MIME type, extension, and size for all uploads",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "File Storage Integration",
     "file-storage.ts integrates with S3-compatible storage for all file uploads",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "MongoDB Transaction Management",
     "Multi-document transactions in exchange and profile operations via clientSession",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Input Validation (Zod Schemas)",
     "schemas.ts defines Zod schemas for all inputs; validated in every server action",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "CORS Policy Enforcement",
     "CORS headers set on API routes restricting to whitelisted origins",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Structured Error Logging",
     "error-handling.ts + logger.ts write structured server-side error logs with context",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Environment Variable Validation",
     "env-validation.ts validates required env vars on startup and prevents silent failures",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Business Logic Middleware",
     "business-logic-middleware.ts enforces ownership and eligibility before handlers",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Resource Authorization",
     "resource-authorization.ts / validateResourceAccess centralizes permission enforcement",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Server Action Authorization",
     "withAuthenticatedAction / withAuthenticatedUserFull verify session on every mutation",
     "System", "Yes", "No", "No", "High"),

    ("Security & Infrastructure", "Database Maintenance & Health Checks",
     "database-maintenance.ts runDatabaseMaintenance covers index creation, TTL, cleanup",
     "System", "Yes", "No", "No", "Medium"),
]

# ── Style definitions ────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)

MODULE_COLORS = {
    "Authentication":           "D6E4F0",
    "User Profile":             "D5F5E3",
    "Book":                     "FEF9E7",
    "Exchange":                 "FDEDEC",
    "Community":                "F4ECF7",
    "Messaging":                "EAF4FB",
    "Notifications":            "FDFEFE",
    "Admin Dashboard":          "FBEEE6",
    "AI":                       "EBF5FB",
    "Security & Infrastructure":"F2F3F4",
}

PRIORITY_COLORS = {
    "High":   "E74C3C",
    "Medium": "F39C12",
    "Low":    "27AE60",
}

BOOL_COLORS = {
    "Yes": "27AE60",
    "No":  "95A5A6",
}

# ── Write headers ────────────────────────────────────────────────────────────
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font   = HEADER_FONT
    cell.fill   = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER

# ── Write data rows ──────────────────────────────────────────────────────────
for row_idx, row_data in enumerate(features, start=2):
    module   = row_data[0]
    priority = row_data[7]

    bg_hex   = MODULE_COLORS.get(module, "FFFFFF")
    bg_fill  = PatternFill("solid", fgColor=bg_hex)

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.fill = bg_fill
        cell.font = Font(size=10, name="Calibri")

        # Colour-code Priority column
        if col_idx == 8:
            colour = PRIORITY_COLORS.get(value, "000000")
            cell.font = Font(bold=True, color=colour, size=10, name="Calibri")

        # Colour-code boolean columns (5, 6, 7)
        if col_idx in (5, 6, 7):
            colour = BOOL_COLORS.get(value, "000000")
            cell.font = Font(bold=True, color=colour, size=10, name="Calibri")

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = [28, 38, 70, 40, 18, 18, 14, 12]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Row height ───────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36
for row_idx in range(2, len(features) + 2):
    ws.row_dimensions[row_idx].height = 40

# ── Freeze header row ────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Auto-filter ──────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── Summary sheet ────────────────────────────────────────────────────────────
ws_summary = wb.create_sheet(title="Module Summary")

from collections import Counter
module_counts   = Counter(f[0] for f in features)
priority_counts = Counter(f[7] for f in features)

# Summary header
for col, txt in enumerate(["Module", "Feature Count"], start=1):
    c = ws_summary.cell(row=1, column=col, value=txt)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = THIN_BORDER

ws_summary.column_dimensions["A"].width = 35
ws_summary.column_dimensions["B"].width = 18
ws_summary.row_dimensions[1].height = 30

for r, (mod, cnt) in enumerate(sorted(module_counts.items()), start=2):
    bg = PatternFill("solid", fgColor=MODULE_COLORS.get(mod, "FFFFFF"))
    c1 = ws_summary.cell(row=r, column=1, value=mod)
    c2 = ws_summary.cell(row=r, column=2, value=cnt)
    for c in (c1, c2):
        c.fill = bg
        c.border = THIN_BORDER
        c.font = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
    c1.alignment = Alignment(horizontal="left", vertical="center")

# Total row
total_row = len(module_counts) + 2
c_total   = ws_summary.cell(row=total_row, column=1, value="TOTAL")
c_count   = ws_summary.cell(row=total_row, column=2, value=len(features))
for c in (c_total, c_count):
    c.font   = Font(bold=True, size=11, name="Calibri")
    c.fill   = PatternFill("solid", fgColor="1F3864")
    c.font   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

# Priority breakdown
ws_summary.cell(row=total_row + 2, column=1,
                value="Priority Breakdown").font = Font(bold=True, size=11, name="Calibri")
for col, txt in enumerate(["Priority", "Count"], start=1):
    c = ws_summary.cell(row=total_row + 3, column=col, value=txt)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = THIN_BORDER
    c.alignment = HEADER_ALIGN

for r, p in enumerate(["High", "Medium", "Low"], start=total_row + 4):
    c1 = ws_summary.cell(row=r, column=1, value=p)
    c2 = ws_summary.cell(row=r, column=2, value=priority_counts.get(p, 0))
    colour = PRIORITY_COLORS.get(p, "000000")
    for c in (c1, c2):
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=True, color=colour, size=10, name="Calibri")

# ── Save ─────────────────────────────────────────────────────────────────────
output_path = "/Users/sabihhaider/Documents/BookEX/BookEx_Feature_Inventory.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Total features: {len(features)}")
